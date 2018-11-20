from django.contrib.auth.hashers import make_password, check_password
from models.models import *
from io import BytesIO
from openpyxl import load_workbook, workbook, cell
from .forms import CreateGroup
from django.contrib.auth.models import User
from django.db import transaction
import operator
from django.db.models import Q
from main.stageManagement import createNewStudentStages
from .absenceManagement import *

def createStudentFromWorkbook(file_in_memory, group):
    try:
        wb = load_workbook(filename=BytesIO(file_in_memory))
        for sheetname in wb.sheetnames:
            rows = wb[sheetname].max_row
            columns = wb[sheetname].max_column
            ws = wb[sheetname]
            firstname_index = 2
            lastname_index = 3
            index_number_index = 4
            for i in range(1,columns+1):
                cell_value = ws.cell(row=1,column=i).value
                if cell_value == "Imię" or cell_value == "Firstname":
                    firstname_index = i
                elif cell_value == "Nazwisko" or cell_value == "Lastname":
                    lastname_index = i
                elif cell_value == "Numer Indeksu" or cell_value == "Index" or cell_value == "Nr_Indeksu":
                    index_number_index = i
            for i in range (2, rows+1):
                index_nr = str(ws.cell(row=i,column=index_number_index).value)
                username = 's'+str(ws.cell(row=i,column=index_number_index).value)
                if not User.objects.filter(username=username).exists():
                    first_name = ws.cell(row=i,column=firstname_index).value
                    last_name = ws.cell(row=i,column=lastname_index).value 
                    user = createStudent(first_name,last_name,index_nr)
                    studentGroup = StudentGroup.objects.create(group=group, student=user)
                    studentGroup.save()
        return True
    except:
        return False

def createStudent(first_name, last_name, index_nr):
    try:
        password = first_name[:3] + last_name[:3] + str(index_nr)[:3]
        username = 's'+str(index_nr)
        if User.objects.filter(username=username).exists():
            return False
        user = User.objects.create_user(username=username,password=password, first_name=first_name, last_name=last_name)
        accountDetails = AccountDetails.objects.create(
            user=user,
            level=0,
        )
        createAbsenceList(user)
        createNewStudentStages(user)
        return user
    except:
        return None

def getUser(user_id):
    if User.objects.filter(id=user_id).exists():
        return User.objects.filter(id=user_id).first()

def deleteStudent(student):
    deletePermamentMessageWithAnswers()
    User.objects.filter(is_superuser=False).filter(id=student.id).delete()

def getAllStudentsFromGroup(group_id):
    results =[]
    group = Group.objects.filter(id=group_id).first()
    studentGroupList = StudentGroup.objects.filter(group = group)
    for studentGroup in studentGroupList:
        results.append(studentGroup.student.id)
    return results

def setUserAbsence(post, group_id):
    studentsList = getAllStudentsFromGroup(group_id)
    for student in studentsList:
        absenceList = getAllStudentAbsence(getUser(student))
        for absence in absenceList:
            post_id = "lesson_"+str(student)+"__"+str(absence.id)
            if post_id in post:
                setAbsenceForStudent(absence.id)
    return True
